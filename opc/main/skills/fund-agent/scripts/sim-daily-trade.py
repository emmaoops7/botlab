#!/usr/bin/env python3
"""
模拟盘每日交易 — 选股 + 决策 + 执行
每日运行：自动选股 → 买入信号 → 持仓检查
"""
import os
import sys
import json
import subprocess
from pathlib import Path
from datetime import datetime

DATA_DIR = Path(__file__).parent.parent / "data"
TRADE_LOG = DATA_DIR / "sim-trades.json"

MX_APIKEY = os.environ.get("MX_APIKEY", "")
MX_XUANGU = None  # mx-xuangu not configured
MX_MONI = None  # mx-moni not configured
MX_DATA = None  # mx-data not configured

SIM_CAPITAL = 30000  # 模拟本金上限 3 万元


def run_mx(script, query, timeout=30):
    if not MX_APIKEY:
        return {"error": "MX_APIKEY missing"}
    env = {**os.environ, "MX_APIKEY": MX_APIKEY}
    result = subprocess.run(
        [sys.executable, str(script), query],
        capture_output=True, text=True, timeout=timeout, env=env
    )
    return result


def get_account():
    r = run_mx(MX_MONI, "我的资金")
    lines = r.stdout.strip().split("\n")
    total = 0
    available = 0
    for line in lines:
        if "总资产" in line:
            try:
                total = float(line.split(":")[1].strip().replace("元", "").replace(",", "").strip())
            except:
                pass
        if "可用资金" in line:
            try:
                available = float(line.split(":")[1].strip().replace("元", "").replace(",", "").strip())
            except:
                pass
    return total, available


def get_positions():
    r = run_mx(MX_MONI, "我的持仓")
    stocks = []
    for line in r.stdout.split("\n"):
        parts = line.strip().split()
        if len(parts) >= 4 and parts[0].isdigit() and len(parts[0]) == 6:
            stocks.append({
                "code": parts[0],
                "name": parts[1] if len(parts) > 1 else "",
                "qty": int(parts[-2]) if parts[-2].isdigit() else 0,
                "price": float(parts[-1]) if parts[-1].replace(".", "").isdigit() else 0,
            })
    return stocks


def screen_stocks():
    """选股：市盈率<30，净利润增长>10%，市值>100亿"""
    r = run_mx(MX_XUANGU, "市盈率小于30 净利润增长率大于10 流通市值大于100亿")
    if r.returncode != 0:
        return []
    csv_path = None
    for line in r.stdout.split("\n"):
        if "CSV" in line:
            csv_path = line.split("CSV: ")[-1].strip()
            break
    if not csv_path or not Path(csv_path).exists():
        return []
    try:
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            lines = f.readlines()
        if len(lines) < 2:
            return []
        stocks = []
        for line in lines[1:6]:  # 前5只
            parts = line.strip().split(",")
            if len(parts) >= 5:
                stocks.append({
                    "code": parts[1],      # 代码
                    "name": parts[2],      # 名称
                    "price": parts[4],     # 最新价
                })
        return stocks
    except Exception as e:
        print(f"  CSV解析失败: {e}")
        return []


def get_stock_price(code):
    """从 mx-data 获取股票最新价（返回 Excel）"""
    r = run_mx(MX_DATA, f"{code} 实时行情")
    if r.returncode != 0:
        return 0
    xlsx_path = None
    for line in r.stdout.split("\n"):
        if ".xlsx" in line:
            xlsx_path = line.split(".xlsx")[0].strip().split(": ")[-1].strip() + ".xlsx"
            break
    if not xlsx_path or not Path(xlsx_path).exists():
        return 0
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet_name in wb.sheetnames:
            if "涨跌幅" in sheet_name:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 2:
                    headers = [str(x) for x in rows[0]]
                    for i, h in enumerate(headers):
                        if "最新价" in h:
                            val = rows[1][i] if len(rows[1]) > i else None
                            if val:
                                return float(val)
    except Exception as e:
        print(f"  Excel解析: {e}")
    return 0


def check_5day_rise(code):
    """检查5日涨幅，超过10%不买（高位不追）"""
    r = run_mx(MX_DATA, f"{code} 实时行情")
    if r.returncode != 0:
        return 0
    xlsx_path = None
    for line in r.stdout.split("\n"):
        if ".xlsx" in line:
            xlsx_path = line.split(".xlsx")[0].strip().split(": ")[-1].strip() + ".xlsx"
            break
    if not xlsx_path or not Path(xlsx_path).exists():
        return 0
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        for sheet_name in wb.sheetnames:
            if "涨跌幅" in sheet_name:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) >= 2:
                    headers = [str(x) for x in rows[0]]
                    for i, h in enumerate(headers):
                        if "5日涨跌幅" in h:
                            val = rows[1][i] if len(rows[1]) > i else None
                            if val:
                                return float(str(val).replace("%", ""))
    except:
        pass
    return 0


def check_stop_loss(code, name, buy_price, current_price, threshold=0.05):
    """止损检查：跌幅超过阈值返回 True"""
    if buy_price <= 0 or current_price <= 0:
        return False
    loss = (current_price - buy_price) / buy_price
    if loss <= -threshold:
        return True
    return False


def execute_buy(code, name, price_round=2):
    """买入：用可用资金的 20%，且不超过模拟本金上限，避开高位股"""
    total, available = get_account()
    sim_available = min(available, SIM_CAPITAL)
    if sim_available < 1000:
        print(f"  资金不足 ¥{sim_available:,.0f}")
        return

    # 检查5日涨幅，>10%不买
    rise_5d = check_5day_rise(code)
    if rise_5d > 10:
        print(f"  ⚠️ {name}({code}) 5日涨{rise_5d}%，高位不追")
        return

    price = get_stock_price(code)
    if price <= 0:
        print(f"  无法获取 {name}({code}) 价格，跳过")
        return

    # 计算买入数量（可用资金的 20%，向下取 100 的倍数）
    budget = sim_available * 0.20
    qty = int(budget / price / 100) * 100
    if qty < 100:
        print(f"  {name}({code}) 最新 ¥{price}，可买不足 100 股")
        return

    # 市价买入
    r = run_mx(MX_MONI, f"市价买入 {code} {qty} 股")
    if "成功" in r.stdout:
        print(f"  ✅ 买入 {name}({code}) {qty}股 @ ~¥{price}")
        log_trade("buy", code, name, price, qty)
    else:
        print(f"  ❌ 买入失败: {r.stdout[:100]}")


def log_trade(action, code, name, price, qty):
    trades = []
    if TRADE_LOG.exists():
        try:
            with open(TRADE_LOG, encoding="utf-8") as f:
                trades = json.load(f)
        except:
            trades = []
    trades.append({
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        "code": code,
        "name": name,
        "price": price,
        "qty": qty,
        "amount": price * qty,
    })
    with open(TRADE_LOG, "w", encoding="utf-8") as f:
        json.dump(trades, f, ensure_ascii=False, indent=2)


def daily_run():
    print("📊 模拟盘每日交易")
    print("=" * 40)

    # 1. 账户状态（限制为 1 万元模拟本金）
    total, available = get_account()
    sim_total = min(total, SIM_CAPITAL)
    sim_available = min(available, SIM_CAPITAL)
    print(f"💰 模拟本金: ¥{SIM_CAPITAL:,.0f} | 可用: ¥{sim_available:,.0f}")

    # 2. 当前持仓
    positions = get_positions()
    if positions:
        print(f"📦 持仓 ({len(positions)} 只):")
        for p in positions:
            cur_price = get_stock_price(p["code"])
            if cur_price > 0:
                pnl = (cur_price - p["price"]) / p["price"] * 100 if p["price"] > 0 else 0
                icon = "🟢" if pnl >= 0 else "🔴"
                print(f"  {icon} {p['name']}({p['code']}) {p['qty']}股 成本¥{p['price']} 现价¥{cur_price} 盈亏{pnl:+.1f}%")
                # 止损：跌幅超过5%自动卖出
                if pnl <= -5:
                    print(f"  🚨 止损触发！{p['name']} 跌{pnl:.1f}%，执行卖出")
                    r = run_mx(MX_MONI, f"市价卖出 {p['code']} {p['qty']} 股")
                    if "成功" in r.stdout:
                        print(f"  ✅ 止损卖出 {p['name']} {p['qty']}股")
                        log_trade("sell_stop", p["code"], p["name"], cur_price, p["qty"])
                    else:
                        print(f"  ❌ 止损卖出失败: {r.stdout[:100]}")
            else:
                print(f"  {p['name']}({p['code']}) {p['qty']}股")
    else:
        print("📦 空仓")

    # 3. 选股 + 买入信号
    print("\n🎯 选股 + 建仓:")
    candidates = screen_stocks()
    if candidates:
        for c in candidates[:3]:  # 最多买 3 只
            print(f"  考虑: {c['name']}({c['code']})")
            execute_buy(c["code"], c["name"])
    else:
        print("  无符合条件的股票")

    print("\n✅ 模拟盘每日交易完成")


if __name__ == "__main__":
    daily_run()
